import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename: str):
    # Hierarchy: Workbooks -> Worksheet -> Cell -> Value
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    # Correcting cell values
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # This is like highlighting a group of cells
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    # BarChart may have some interesting options for customizing UI
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'f2')

    wb.save(filename)
